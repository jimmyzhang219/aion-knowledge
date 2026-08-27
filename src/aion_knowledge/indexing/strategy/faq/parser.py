"""FAQ 文件解析器：支持 CSV、Excel (.xlsx)、JSON 三种格式。"""

from __future__ import annotations

import csv
import io
import json
import logging
import zipfile
from typing import Any

from aion_knowledge.indexing.strategy.faq.schemas import FAQEntry

logger = logging.getLogger(__name__)


class ParseError(ValueError):
    """FAQ 文件解析错误。"""
    pass


# 字段别名映射：支持中英文表头
_FIELD_ALIASES: dict[str, list[str]] = {
    "standard_question": ["问题", "standard_question", "question", "Q"],
    "answers":           ["答案", "机器人回答", "answers", "answer", "A"],
    "similar_questions": ["相似问题", "similar_questions"],
    "negative_questions": ["负问题", "反例问题", "negative_questions"],
    "category":          ["分类", "category", "tag_name", "tags"],
    "answer_strategy":   ["答案策略", "answer_strategy"],
}


def _get_field(row: dict[str, str], field: str) -> str:
    """按别名优先级从行数据中取字段值。"""
    for alias in _FIELD_ALIASES.get(field, []):
        if alias in row:
            return row[alias].strip()
    return ""


def _detect_delimiter(sample: str) -> str:
    """Detect CSV delimiter (comma or semicolon) from header line."""
    first_line = sample.split("\n", 1)[0]
    comma_count = first_line.count(",")
    semi_count = first_line.count(";")
    return ";" if semi_count > comma_count else ","


def _read_csv(content: str) -> list[dict[str, str]]:
    """Parse CSV content into list of dicts."""
    if not content.strip():
        raise ParseError("CSV 内容为空")

    delimiter = _detect_delimiter(content)

    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    if not reader.fieldnames:
        raise ParseError("CSV 缺少表头")

    rows: list[dict[str, str]] = []
    for row in reader:
        if any(v.strip() for v in row.values() if v):
            rows.append({k.strip(): (v.strip() if v else "") for k, v in row.items()})
    return rows


def _parse_csv_rows(rows: list[dict[str, str]]) -> list[FAQEntry]:
    """Convert parsed CSV rows to FAQEntry list. 支持中英文表头别名。"""
    results: list[FAQEntry] = []
    for row in rows:
        question = _get_field(row, "standard_question")
        answers_raw = _get_field(row, "answers")
        if not question and not answers_raw:
            continue
        if not question:
            raise ParseError(f"行缺少问题: {row}")
        if not answers_raw:
            raise ParseError(f"问题 '{question}' 缺少答案")

        answers = [a.strip() for a in answers_raw.split("##") if a.strip()]
        similar = [s.strip() for s in _get_field(row, "similar_questions").split("##") if s.strip()]
        negative = [n.strip() for n in _get_field(row, "negative_questions").split("##") if n.strip()]
        tag_raw = _get_field(row, "category")
        strategy = _get_field(row, "answer_strategy") or "all"
        _validate_entry(question, similar, negative, answers)

        results.append(FAQEntry(
            standard_question=question,
            similar_questions=similar,
            negative_questions=negative,
            answers=answers,
            answer_strategy=strategy if strategy in ("all", "random") else "all",
            tags=[tag_raw] if tag_raw else [],
        ))
    return results


def _validate_entry(
    question: str, similar: list[str], negative: list[str], answers: list[str],
) -> None:
    """Validate a single FAQ entry. Raises ParseError on violation."""
    if not question.strip():
        raise ParseError("standard_question 不能为空")
    if not answers:
        raise ParseError(f"问题 '{question}' 必须至少有一个答案")
    for sq in similar:
        if sq.strip().lower() == question.strip().lower():
            raise ParseError(f"相似问题 '{sq}' 不能与标准问题 '{question}' 相同")
    for nq in negative:
        if nq.strip().lower() == question.strip().lower():
            raise ParseError(f"负问题 '{nq}' 不能与标准问题 '{question}' 相同")
        for sq in similar:
            if nq.strip().lower() == sq.strip().lower():
                raise ParseError(f"负问题 '{nq}' 不能与相似问题 '{sq}' 相同")


def _parse_json(content: str) -> list[FAQEntry]:
    """Parse JSON content into FAQEntry list."""
    try:
        data: list[dict[str, Any]] = json.loads(content)
    except json.JSONDecodeError as e:
        raise ParseError(f"JSON 解析失败: {e}")

    if not isinstance(data, list):
        raise ParseError("JSON 必须是数组")

    results: list[FAQEntry] = []
    for entry in data:
        question = (entry.get("standard_question") or "").strip()
        answers = entry.get("answers", [])
        similar = [s.strip() for s in entry.get("similar_questions", []) if s.strip()]
        negative = [n.strip() for n in entry.get("negative_questions", []) if n.strip()]
        strategy = entry.get("answer_strategy", "all")
        tags = entry.get("tags", [])
        _validate_entry(question, similar, negative, answers)

        results.append(FAQEntry(
            standard_question=question,
            similar_questions=similar,
            negative_questions=negative,
            answers=answers,
            answer_strategy=strategy if strategy in ("all", "random") else "all",
            tags=tags if isinstance(tags, list) else [],
        ))
    return results


def _read_excel(content: bytes) -> list[dict[str, str]]:
    """Parse Excel .xlsx content into list of dicts."""
    try:
        import openpyxl  # type: ignore[import-untyped]  # noqa: F401  # openpyxl 未安装 stub（types-openpyxl）
    except ImportError:
        raise ParseError("openpyxl 未安装，无法解析 .xlsx 文件")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except zipfile.BadZipFile:
        raise ParseError("文件格式错误：无法解析为 Excel 文件")
    ws = wb.active
    if ws is None:
        raise ParseError("Excel 文件没有工作表")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ParseError("Excel 文件为空")

    header = [str(c).strip() if c is not None else "" for c in header]
    rows: list[dict[str, str]] = []
    for row in rows_iter:
        row_dict = {}
        has_content = False
        for i, val in enumerate(row):
            key = header[i] if i < len(header) else ""
            raw = str(val).strip() if val is not None else ""
            row_dict[key] = raw
            if raw:
                has_content = True
        if has_content:
            rows.append(row_dict)
    return rows


def parse_faq_file(content: bytes, file_ext: str) -> list[FAQEntry]:
    """解析 FAQ 文件内容为 FAQEntry 列表。"""
    ext = file_ext.lower().lstrip(".")

    if ext == "json":
        return _parse_json(content.decode("utf-8-sig"))
    elif ext in ("csv",):
        rows = _read_csv(content.decode("utf-8-sig"))
        return _parse_csv_rows(rows)
    elif ext in ("xlsx", "xls"):
        rows = _read_excel(content)
        return _parse_csv_rows(rows)
    else:
        raise ParseError(f"不支持的文件格式: {file_ext}")
