import io
import os
import shutil
import subprocess
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

from aion_knowledge.pipeline.parser.office import (
    ExcelParser,
    convert_ppt_to_pptx_bytes,
    detect_excel_format,
    engine_for_format,
    fill_merged_cells_xlsx,
    is_ole_compound,
    is_zip_openxml,
    needs_ppt_to_pptx_conversion,
    normalize_ppt_bytes,
    repair_xlsx_bytes,
)

TESTDATA = Path(__file__).resolve().parents[4] / "testdata" / "rag_test"
LEGACY_PPT = TESTDATA / "ppt_old" / "en_38256.ppt"
WMF_IMAGE_PPT = LEGACY_PPT
IMAGE_HEAVY_PPT = TESTDATA / "ppt_old" / "en_41384.ppt"
PPTX_SAMPLE = TESTDATA / "pptx" / "en_marker.pptx"


# ── Excel Tests ────────────────────────────────────────────────────────────

def _xlsx_with_phantom_shared_strings() -> bytes:
    """Workbook with inline strings but a dangling sharedStrings manifest entry."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B1"] = 42
    bio = io.BytesIO()
    wb.save(bio)

    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(io.BytesIO(bio.getvalue()), "r") as zin:
            zin.extractall(tmpdir)

        ct_path = f"{tmpdir}/[Content_Types].xml"
        with open(ct_path, encoding="utf-8") as f:
            ct = f.read()
        override = (
            '<Override PartName="/xl/sharedStrings.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sharedStrings+xml"/>'
        )
        with open(ct_path, "w", encoding="utf-8") as f:
            f.write(ct.replace("</Types>", override + "</Types>"))

        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for root, _, files in os.walk(tmpdir):
                for name in files:
                    path = os.path.join(root, name)
                    arc = os.path.relpath(path, tmpdir)
                    zout.write(path, arc)
        return out.getvalue()


class ExcelFormatDetectionTest(unittest.TestCase):
    def test_detect_xlsx_and_engine(self):
        wb = openpyxl.Workbook()
        bio = io.BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        self.assertEqual(detect_excel_format(content), "xlsx")
        self.assertEqual(engine_for_format("xlsx"), "openpyxl")

    def test_detect_xls_magic(self):
        content = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512
        self.assertEqual(detect_excel_format(content), "xls")
        self.assertEqual(engine_for_format("xls"), "xlrd")

    def test_open_legacy_xls_bytes_with_xlsx_extension(self):
        if not shutil.which("soffice"):
            self.skipTest("LibreOffice not available")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "legacy"
        xlsx_bio = io.BytesIO()
        wb.save(xlsx_bio)
        with tempfile.TemporaryDirectory() as tmpdir:
            src = os.path.join(tmpdir, "sheet.xlsx")
            with open(src, "wb") as handle:
                handle.write(xlsx_bio.getvalue())
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "xls",
                    "--outdir",
                    tmpdir,
                    src,
                ],
                check=True,
                capture_output=True,
            )
            xls_path = os.path.join(tmpdir, "sheet.xls")
            with open(xls_path, "rb") as handle:
                xls_bytes = handle.read()

        document = ExcelParser(file_name="fake.xlsx", file_type="xlsx").parse_into_text(
            xls_bytes
        )
        self.assertIn("legacy", document.content)


class XlsxRepairTest(unittest.TestCase):
    def test_repair_removes_phantom_shared_strings_reference(self):
        broken = _xlsx_with_phantom_shared_strings()
        with self.assertRaises(KeyError):
            pd.read_excel(io.BytesIO(broken))

        repaired = repair_xlsx_bytes(broken)
        self.assertIsNotNone(repaired)
        df = pd.read_excel(io.BytesIO(repaired), header=None)
        self.assertEqual(df.values.tolist(), [["hello", 42]])

    def test_repair_skips_when_shared_string_cells_need_table(self):
        try:
            import xlsxwriter
        except ImportError:
            self.skipTest("xlsxwriter not available")

        bio = io.BytesIO()
        wb = xlsxwriter.Workbook(bio, {"in_memory": True})
        ws = wb.add_worksheet()
        ws.write(0, 0, "hello")
        wb.close()

        with tempfile.TemporaryDirectory() as tmpdir:
            with zipfile.ZipFile(io.BytesIO(bio.getvalue()), "r") as zin:
                zin.extractall(tmpdir)
            os.remove(f"{tmpdir}/xl/sharedStrings.xml")

            out = io.BytesIO()
            with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
                for root, _, files in os.walk(tmpdir):
                    for name in files:
                        path = os.path.join(root, name)
                        arc = os.path.relpath(path, tmpdir)
                        zout.write(path, arc)
            broken = out.getvalue()

        self.assertIsNone(repair_xlsx_bytes(broken))


class XlsxMergeFillTest(unittest.TestCase):
    def test_fill_merged_cells_propagates_master_value(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "title"
        ws.merge_cells("A1:B1")
        ws["A2"] = "left"
        ws["B2"] = "right"
        ws.merge_cells("A2:A3")
        ws["B3"] = "only-b"
        bio = io.BytesIO()
        wb.save(bio)

        filled = fill_merged_cells_xlsx(bio.getvalue())
        out_wb = openpyxl.load_workbook(io.BytesIO(filled), data_only=True)
        out_ws = out_wb.active
        self.assertEqual(out_ws["B1"].value, "title")
        self.assertEqual(out_ws["A3"].value, "left")
        self.assertEqual(out_ws["B3"].value, "only-b")

    def test_parse_en_mergecell_workbook(self):
        path = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "..",
            "..",
            "testdata",
            "rag_test",
            "xlsx",
            "en_mergecell.xlsx",
        )
        if not os.path.isfile(path):
            self.skipTest("en_mergecell.xlsx fixture not available")
        with open(path, "rb") as handle:
            document = ExcelParser().parse_into_text(handle.read())

        self.assertNotIn("Unnamed:", document.content)
        self.assertIn("A: A1", document.content)
        self.assertIn("D: D10", document.content)


class ExcelImageFilterTest(unittest.TestCase):
    """Tests for filtering embedded image function strings (#1779)."""

    def _xlsx_with_image_functions(self) -> bytes:
        """Create an XLSX where image functions are stored as text values.

        WPS embeds images using =DISPIMG("ID",1) which may appear as plain
        text (not a formula) in some export scenarios.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Photo"
        ws["A2"] = "Alice"
        ws["B2"] = '_xlfn.DISPIMG("ID_ABCDEF123",1)'
        ws["A3"] = "Bob"
        ws["B3"] = '_xlfn.DISPIMG("ID_GHIJKL456",1)'
        ws["A4"] = "Charlie"
        ws["B4"] = "real data"
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def test_dispimg_text_values_are_excluded(self):
        """Image function strings stored as text must not appear in output."""
        document = ExcelParser().parse_into_text(self._xlsx_with_image_functions())
        self.assertNotIn("DISPIMG", document.content)
        self.assertNotIn("_xlfn", document.content)
        # Real data must still be present
        self.assertIn("Alice", document.content)
        self.assertIn("Bob", document.content)
        self.assertIn("Charlie", document.content)
        self.assertIn("real data", document.content)

    def test_dispimg_with_equals_prefix(self):
        """=_xlfn.DISPIMG(...) as text (not formula) should also be filtered."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Photo"
        ws["A2"] = "Alice"
        # Stored as text with = prefix (some exporters do this)
        ws["B2"] = '=_xlfn.DISPIMG("ID_123",1)'
        bio = io.BytesIO()
        wb.save(bio)
        # Note: openpyxl treats strings starting with = as formulas,
        # so data_only=True in fill_merged_cells_xlsx will turn them to None.
        # This test verifies the formula path also works correctly.
        document = ExcelParser().parse_into_text(bio.getvalue())
        self.assertNotIn("DISPIMG", document.content)
        self.assertIn("Alice", document.content)

    def test_image_function_variations(self):
        """Various image function patterns should all be filtered."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = '_xlfn.DISPIMG("ID_001",1)'
        ws["A2"] = 'DISPIMG("ID_002",1)'  # no prefix at all
        ws["A3"] = '_xlfn.IMAGE("https://example.com/img.png")'
        ws["A4"] = 'IMAGE("https://example.com/img.png",1)'
        ws["A5"] = "Normal text"
        bio = io.BytesIO()
        wb.save(bio)
        document = ExcelParser().parse_into_text(bio.getvalue())
        self.assertNotIn("DISPIMG", document.content)
        self.assertNotIn("IMAGE(", document.content)
        self.assertIn("Normal text", document.content)

    def test_real_world_wps_dispimg(self):
        """Exact pattern from issue #1779 screenshot: =DISPIMG("ID_...",1)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Product"
        ws["B1"] = "Image"
        ws["C1"] = "Price"
        ws["A2"] = "Basic"
        # This is the exact format shown in the issue screenshot
        ws["B2"] = 'DISPIMG("ID_5A60F9ED501E48A38EBEE5D326E18235",1)'
        ws["C2"] = "39.9"
        ws["A3"] = "Pro"
        ws["B3"] = 'DISPIMG("ID_AABBCCDD",1)'
        ws["C3"] = "99.9"
        bio = io.BytesIO()
        wb.save(bio)
        document = ExcelParser().parse_into_text(bio.getvalue())
        self.assertNotIn("DISPIMG", document.content)
        self.assertNotIn("ID_5A60F9ED", document.content)
        self.assertIn("Basic", document.content)
        self.assertIn("39.9", document.content)
        self.assertIn("Pro", document.content)
        self.assertIn("99.9", document.content)


class ExcelParserTest(unittest.TestCase):
    def test_parse_phantom_shared_strings_workbook(self):
        document = ExcelParser().parse_into_text(_xlsx_with_phantom_shared_strings())
        self.assertIn("hello", document.content)
        self.assertIn("42", document.content)

    def test_parse_en_calcchain_shared_strings_case(self):
        path = os.path.join(
            os.path.dirname(__file__),
            "..",
            "..",
            "..",
            "..",
            "testdata",
            "rag_test",
            "xlsx",
            "en_calcchain.xlsx",
        )
        if not os.path.isfile(path):
            self.skipTest("en_calcchain.xlsx fixture not available")
        with open(path, "rb") as f:
            document = ExcelParser().parse_into_text(f.read())
        self.assertGreater(len(document.content), 0)


# ── PPT Tests ──────────────────────────────────────────────────────────────

class TestPptConvert(unittest.TestCase):
    def test_legacy_ppt_magic(self):
        if not LEGACY_PPT.is_file():
            self.skipTest("testdata missing")
        content = LEGACY_PPT.read_bytes()
        self.assertTrue(is_ole_compound(content))
        self.assertFalse(is_zip_openxml(content))
        self.assertTrue(needs_ppt_to_pptx_conversion(content, "ppt"))

    def test_pptx_does_not_need_conversion(self):
        if not PPTX_SAMPLE.is_file():
            self.skipTest("testdata missing")
        content = PPTX_SAMPLE.read_bytes()
        self.assertTrue(is_zip_openxml(content))
        self.assertFalse(needs_ppt_to_pptx_conversion(content, "pptx"))

    def test_normalize_pptx_passthrough(self):
        if not PPTX_SAMPLE.is_file():
            self.skipTest("testdata missing")
        content = PPTX_SAMPLE.read_bytes()
        out, ext = normalize_ppt_bytes(content, "pptx")
        self.assertEqual(out, content)
        self.assertEqual(ext, ".pptx")

    def test_legacy_ppt_requires_soffice(self):
        if not LEGACY_PPT.is_file():
            self.skipTest("testdata missing")
        if not shutil.which("soffice"):
            with self.assertRaises(ValueError) as ctx:
                normalize_ppt_bytes(LEGACY_PPT.read_bytes(), "ppt")
            self.assertIn("LibreOffice", str(ctx.exception))
            self.skipTest("LibreOffice not available")
        converted = convert_ppt_to_pptx_bytes(LEGACY_PPT.read_bytes(), suffix=".ppt")
        self.assertIsNotNone(converted)
        self.assertTrue(is_zip_openxml(converted))
        out, ext = normalize_ppt_bytes(LEGACY_PPT.read_bytes(), "ppt")
        self.assertEqual(ext, ".pptx")
        self.assertTrue(is_zip_openxml(out))

    def test_wmf_legacy_ppt_extracts_rasterized_image(self):
        if not shutil.which("soffice"):
            self.skipTest("LibreOffice not available")
        if not shutil.which("convert"):
            self.skipTest("ImageMagick convert not available")
        if not WMF_IMAGE_PPT.is_file():
            self.skipTest("testdata missing")

        from aion_knowledge.pipeline.parser.markitdown import MarkitdownParser

        doc = MarkitdownParser(file_type="ppt").parse_into_text(
            WMF_IMAGE_PPT.read_bytes()
        )
        self.assertEqual(len(doc.images), 1)
        self.assertNotIn("bd10496_.jpg", doc.content)
        self.assertIn("images/", doc.content)

    def test_image_heavy_legacy_ppt_extracts_images(self):
        if not shutil.which("soffice"):
            self.skipTest("LibreOffice not available")
        if not IMAGE_HEAVY_PPT.is_file():
            self.skipTest("testdata missing")

        from aion_knowledge.pipeline.parser.markitdown import MarkitdownParser

        doc = MarkitdownParser(file_type="ppt").parse_into_text(
            IMAGE_HEAVY_PPT.read_bytes()
        )
        self.assertGreaterEqual(len(doc.images), 2)
        self.assertNotIn("![](.jpg)", doc.content)
        for ref in doc.images:
            self.assertTrue(ref.startswith("images/"))


if __name__ == "__main__":
    unittest.main()
